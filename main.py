"""
BotShortTrade – Hyperliquid DEX
Stratégie : M15 biais → M5 confirm → M1 entrée | SL=1.5×ATR | TP=2.5×ATR | 5× levier
"""
import json, math, os, time, traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import requests
from dotenv import load_dotenv
from eth_account import Account
from hyperliquid.info import Info
from hyperliquid.exchange import Exchange
from hyperliquid.utils import constants
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Couleurs console (ANSI) ────────────────────────────────────────────────────
os.system("")  # active les codes ANSI sur Windows

class C:
    RST  = "\033[0m"
    BOLD = "\033[1m"
    DIM  = "\033[2m"
    RED  = "\033[91m"
    GRN  = "\033[92m"
    YLW  = "\033[93m"
    BLU  = "\033[94m"
    CYN  = "\033[96m"
    WHT  = "\033[97m"
    BRED = "\033[1;91m"
    BGRN = "\033[1;92m"
    BYLW = "\033[1;93m"
    BCYN = "\033[1;96m"

def _cp(val: float) -> str:
    """Couleur selon signe d'une valeur P&L."""
    return C.BGRN if val > 0 else (C.BRED if val < 0 else C.WHT)

# ── Env ────────────────────────────────────────────────────────────────────────
load_dotenv()

HL_PRIVATE_KEY    = os.environ["HL_PRIVATE_KEY"]
HL_TESTNET        = os.getenv("HL_TESTNET", "1") == "1"
_MAX_LIQ_CFG  = float(os.getenv("MAX_LIQUIDITY", 0))  # 0 = auto (solde réel)
CAPITAL_PCT       = float(os.getenv("CAPITAL_PCT", 0.10))
LEVERAGE          = int(os.getenv("LEVERAGE", 5))
MIN_MOVE_PCT      = float(os.getenv("MIN_MOVE_PCT", 0.0002))
VOL_MULTIPLIER    = float(os.getenv("VOL_MULTIPLIER", 1.0))
MAX_TRADES        = int(os.getenv("MAX_TRADES", 0))
MAX_DAILY_LOSS_PCT = float(os.getenv("MAX_DAILY_LOSS_PCT", 3.0))   # % du capital
SESSION_START_UTC  = int(os.getenv("SESSION_START_UTC", 0))         # heure UTC (0 = minuit)
SESSION_END_UTC    = int(os.getenv("SESSION_END_UTC", 24))          # heure UTC (24 = H24)
DISCORD_WEBHOOK_URL = os.getenv("DISCORD_WEBHOOK_URL", "")          # webhook Discord (alertes anomalies)

# Initialisées au démarrage dans run() depuis le solde réel
MAX_LIQUIDITY     = _MAX_LIQ_CFG if _MAX_LIQ_CFG > 0 else 100.0  # placeholder
CAPITAL_PER_TRADE = MAX_LIQUIDITY * CAPITAL_PCT

ATR_SL_MULT   = 1.5
ATR_TP_MULT   = 4.0   # R:R ~2.7:1 (TP = 4×ATR, SL = 1.5×ATR)
TRAIL_TRIGGER = 1.5   # activer trailing quand profit ≥ 1.5×ATR
TRAIL_DIST    = 1.0   # trailing SL à 1×ATR du prix courant

# ── Connexion Hyperliquid ──────────────────────────────────────────────────────
_account       = Account.from_key(HL_PRIVATE_KEY)
WALLET_ADDRESS = _account.address
API_URL        = constants.TESTNET_API_URL if HL_TESTNET else constants.MAINNET_API_URL
info           = Info(API_URL, skip_ws=True)
exchange       = Exchange(_account, API_URL)

print(f"  {C.DIM}Wallet : {WALLET_ADDRESS}{C.RST}")
_net_color = C.BYLW if HL_TESTNET else C.BRED
print(f"  Réseau : {_net_color}{'TESTNET' if HL_TESTNET else 'MAINNET'}{C.RST}")

# ── Symboles ───────────────────────────────────────────────────────────────────
WATCHLIST = ["BTC", "ETH", "LTC", "DOGE", "XRP", "SOL"]
SYM_LABEL   = {c: f"{c}/USDT" for c in WATCHLIST}
SYM_BYBIT   = {c: f"{c}USDT" for c in WATCHLIST}   # format Bybit

INTERVAL_M1  = "1m"
INTERVAL_M5  = "5m"
INTERVAL_M15 = "15m"
INTERVAL_D1  = "1D"
_BYBIT_IV    = {"1m": "1", "5m": "5", "15m": "15", "1D": "D"}  # Bybit interval codes

BYBIT_KLINE  = "https://api.bybit.com/v5/market/kline"

# ── Filtres instruments ────────────────────────────────────────────────────────
_sym_filters: dict[str, dict] = {}

def _flt(coin: str) -> dict:
    return _sym_filters.get(coin, {"sz_dec": 3, "sz_step": 0.001})

def init_exchange_info():
    global _sym_filters
    try:
        meta = info.meta()
        for item in meta.get("universe", []):
            coin = item.get("name", "")
            if coin in set(WATCHLIST):
                sz_dec = int(item.get("szDecimals", 3))
                _sym_filters[coin] = {
                    "sz_dec":  sz_dec,
                    "sz_step": 10 ** (-sz_dec),
                    "max_lev": int(item.get("maxLeverage", 50)),
                }
        print(f"  Exchange info : {len(_sym_filters)} symboles chargés")
    except Exception as e:
        log_error(f"init_exchange_info: {e}")

# ── Excel ──────────────────────────────────────────────────────────────────────
EXCEL_FILE = "trades.xlsx"
STATE_FILE = "bot_state.json"
COL_HEADERS = [
    "ID Ordre", "Date/Heure", "Symbole", "Côté", "Qté", "Prix entrée",
    "SL", "TP", "ATR M5", "Notionnel $", "P&L TP $", "P&L SL $",
    "Prix actuel", "Statut", "P&L final $",
]

_FILLS = {
    "header": PatternFill("solid", fgColor="1F3864"),
    "buy":    PatternFill("solid", fgColor="D9EAD3"),
    "sell":   PatternFill("solid", fgColor="FCE5CD"),
    "tp":     PatternFill("solid", fgColor="B6D7A8"),
    "sl":     PatternFill("solid", fgColor="EA9999"),
    "open":   PatternFill("solid", fgColor="FFF2CC"),
}

def _safe_save(wb: Workbook):
    tmp = EXCEL_FILE + ".tmp"
    wb.save(tmp)
    os.replace(tmp, EXCEL_FILE)

def init_excel() -> Workbook:
    if os.path.exists(EXCEL_FILE):
        try:
            return openpyxl.load_workbook(EXCEL_FILE)
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"
    ws.append(COL_HEADERS)
    for c in range(1, len(COL_HEADERS) + 1):
        cell = ws.cell(1, c)
        cell.fill = _FILLS["header"]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    for i in range(3, len(COL_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    _safe_save(wb)
    return wb

def _find_row(ws, order_id: str) -> int | None:
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[0].value) == str(order_id):
            return row[0].row
    return None

def log_order(wb: Workbook, order_id, coin, side, qty, entry, sl, tp, atr, notional, pnl_tp, pnl_sl):
    ws = wb["Trades"]
    fill = _FILLS["buy"] if side == "BUY" else _FILLS["sell"]
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    row = [order_id, ts, SYM_LABEL[coin], side, qty, entry, sl, tp, round(atr, 6),
           notional, pnl_tp, pnl_sl, entry, "Ordre placé", ""]
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.fill = fill
    _safe_save(wb)

def update_order_live(wb: Workbook, order_id, current_px):
    ws = wb["Trades"]
    r = _find_row(ws, order_id)
    if r:
        ws.cell(r, 13).value = round(current_px, 6)
        _safe_save(wb)

def update_order_status(wb: Workbook, order_id, status: str, pnl: float | None = None):
    ws = wb["Trades"]
    r = _find_row(ws, order_id)
    if not r:
        return
    ws.cell(r, 14).value = status
    if pnl is not None:
        ws.cell(r, 15).value = round(pnl, 2)
    fill = _FILLS["tp"] if "TP" in status else (_FILLS["sl"] if "SL" in status else _FILLS["open"])
    for c in range(1, len(COL_HEADERS) + 1):
        ws.cell(r, c).fill = fill
    _safe_save(wb)

# ── Journalisation ────────────────────────────────────────────────────────────
_scan_log:  list[dict] = []
_last_scan: list[dict] = []   # snapshot du dernier scan complet (pour le dashboard)

# ── Apprentissage des erreurs ──────────────────────────────────────────────────
LESSONS_FILE    = "lessons.json"
_current_regime = "RANGING"   # mis à jour dans run() avant chaque scan

def _load_lessons() -> dict:
    try:
        with open(LESSONS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"buckets": {}}

def _save_lessons(data: dict):
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    tmp = LESSONS_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, LESSONS_FILE)

def _lesson_buckets(coin: str, side: str, rsi: float, bbp: float,
                    regime: str, mode: str) -> list[str]:
    """Retourne les clés de bucket statistiques pour un trade donné."""
    s = side.lower()
    # Bucket RSI selon direction et mode
    if s == "buy" and mode == "BBO":
        rsi_tag = "rsi50-62" if rsi < 62 else ("rsi62-70" if rsi < 70 else "rsi70-75")
    elif s == "buy":
        rsi_tag = "rsi<45" if rsi < 45 else ("rsi45-55" if rsi < 55 else "rsi55-70")
    elif mode == "MOM":
        rsi_tag = "rsi<20" if rsi < 20 else "rsi20-35"
    else:  # SELL MR
        rsi_tag = "rsi35-50" if rsi < 50 else "rsi50-65"
    # Bucket BBP selon direction et mode
    if s == "buy" and mode == "BBO":
        bbp_tag = "bbp0.85-1.0" if bbp < 1.0 else "bbp>1.0"
    elif s == "buy":
        bbp_tag = "bbp<0.20" if bbp < 0.20 else "bbp0.20-0.40"
    elif mode == "MOM":
        bbp_tag = "bbp<0.20" if bbp < 0.20 else "bbp0.20-0.35"
    else:  # SELL MR
        bbp_tag = "bbp>0.80" if bbp > 0.80 else "bbp0.65-0.80"
    return [
        f"{rsi_tag}|{s}|{mode}",      # ex: "rsi45-55|buy|MR"
        f"{bbp_tag}|{s}|{mode}",      # ex: "bbp0.20-0.40|buy|MR"
        f"regime_{regime}|{s}|{mode}", # ex: "regime_RANGING|buy|MR"
        f"coin_{coin}|{s}",            # ex: "coin_AVAX|buy"
    ]

def record_lesson(position: dict, outcome: str, pnl: float):
    """Enregistre le résultat d'un trade dans les buckets statistiques."""
    cond = position.get("conditions")
    if not cond:
        return
    data    = _load_lessons()
    buckets = data.setdefault("buckets", {})
    keys    = _lesson_buckets(
        position["symbol"], position["side"],
        cond.get("rsi", 50.0), cond.get("bbp", 0.5),
        cond.get("regime", "RANGING"), cond.get("mode", "MR")
    )
    win = "TP" in outcome or "TRAIL" in outcome
    for k in keys:
        b = buckets.setdefault(k, {"w": 0, "l": 0, "pnl": 0.0})
        if win: b["w"] += 1
        else:   b["l"] += 1
        b["pnl"] = round(b["pnl"] + pnl, 2)
    _save_lessons(data)
    # Afficher une alerte si un bucket devient clairement mauvais
    if not win:
        for k in keys:
            b = buckets[k]
            total = b["w"] + b["l"]
            if total >= 3 and b["w"] / total < 0.35:
                print(f"  {C.BYLW}[LEÇON]{C.RST} Pattern risqué détecté : {C.BOLD}{k}{C.RST} "
                      f"WR={b['w']/total*100:.0f}% ({b['w']}W/{b['l']}L)")

def get_penalty(coin: str, side: str, rsi: float, bbp: float,
                regime: str, mode: str) -> float:
    """Pénalité de score à soustraire si les conditions matchent un historique mauvais."""
    MIN_SAMPLES = 5   # ne pénaliser qu'avec au moins 5 trades dans le bucket
    data    = _load_lessons()
    buckets = data.get("buckets", {})
    keys    = _lesson_buckets(coin, side, rsi, bbp, regime, mode)
    penalty = 0.0
    for k in keys:
        b = buckets.get(k)
        if not b:
            continue
        total = b["w"] + b["l"]
        if total < MIN_SAMPLES:
            continue
        wr = b["w"] / total
        if wr < 0.40:
            penalty += (0.40 - wr) * 0.5   # win_rate=20% → +0.10 | 30% → +0.05
    return round(min(penalty, 0.25), 3)     # plafond : 0.25 max

def print_lessons_summary():
    """Affiche les patterns actuellement pénalisés (si assez de données)."""
    data    = _load_lessons()
    buckets = data.get("buckets", {})
    MIN_SAMPLES = 5
    bad = [(k, v) for k, v in buckets.items()
           if v["w"] + v["l"] >= MIN_SAMPLES and v["w"] / (v["w"] + v["l"]) < 0.40]
    if not bad:
        return
    print(f"\n  {C.BYLW}── Leçons actives ({len(bad)} patterns pénalisés) ──{C.RST}")
    for k, v in sorted(bad, key=lambda x: x[1]["w"] / max(1, x[1]["w"] + x[1]["l"])):
        total = v["w"] + v["l"]
        wr    = v["w"] / total
        pen   = round(min((0.40 - wr) * 0.5, 0.25), 3)
        print(f"    {C.BRED}▸ {k}{C.RST}  WR={wr*100:.0f}% ({v['w']}W/{v['l']}L)  "
              f"P&L={_cp(v['pnl'])}${v['pnl']:+.2f}{C.RST}  pénalité=−{pen}")

def log_error(msg: str):
    ts = datetime.now(timezone.utc).strftime("%H:%M:%S")
    print(f"  {C.BRED}[ERR {ts}]{C.RST} {msg}")

def log_scan(coin: str, tf: str, signal: str | None, score: float, details: str):
    _scan_log.append({"ts": datetime.now(timezone.utc).strftime("%H:%M:%S"),
                      "coin": coin, "tf": tf, "signal": signal,
                      "score": score, "details": details})

# ── Bougies — données Bybit public (sans clé API) ────────────────────────────
def fetch_candles(coin: str, interval: str, count: int, drop_incomplete: bool = True) -> list[dict]:
    symbol = SYM_BYBIT[coin]
    iv     = _BYBIT_IV[interval]
    resp = requests.get(BYBIT_KLINE, params={
        "category": "linear",
        "symbol":   symbol,
        "interval": iv,
        "limit":    count + 2,
    }, timeout=10)
    resp.raise_for_status()
    data = resp.json()
    if data.get("retCode") != 0:
        raise RuntimeError(f"Bybit kline {symbol}: {data.get('retMsg')}")
    # Bybit retourne du plus récent au plus ancien → on inverse
    raw = data["result"]["list"]
    candles = sorted(
        [{"ts": int(r[0]), "o": float(r[1]), "h": float(r[2]),
          "l": float(r[3]), "c": float(r[4]), "v": float(r[5])}
         for r in raw],
        key=lambda x: x["ts"]
    )
    if drop_incomplete and candles:
        candles = candles[:-1]
    return candles[-count:]

# ── Indicateurs ───────────────────────────────────────────────────────────────
def _ema(closes: list[float], n: int) -> list[float]:
    k = 2 / (n + 1)
    ema = [closes[0]]
    for p in closes[1:]:
        ema.append(p * k + ema[-1] * (1 - k))
    return ema

def _atr(candles: list[dict], n: int = 14) -> float:
    trs = []
    for i in range(1, len(candles)):
        c = candles[i]
        pc = candles[i - 1]["c"]
        trs.append(max(c["h"] - c["l"], abs(c["h"] - pc), abs(c["l"] - pc)))
    if not trs:
        return 0.0
    atr = trs[0]
    k = 1 / n
    for tr in trs[1:]:
        atr = tr * k + atr * (1 - k)
    return atr

def _rsi(closes: list[float], n: int = 5) -> float:
    if len(closes) < n + 1:
        return 50.0
    gains, losses = [], []
    for i in range(-n, 0):
        d = closes[i] - closes[i - 1]
        gains.append(max(d, 0))
        losses.append(max(-d, 0))
    ag, al = sum(gains) / n, sum(losses) / n
    if al == 0:
        return 100.0
    return 100 - 100 / (1 + ag / al)

def _vwap(candles: list[dict]) -> float:
    num = sum(((c["h"] + c["l"] + c["c"]) / 3) * c["v"] for c in candles)
    den = sum(c["v"] for c in candles)
    return num / den if den else 0.0

def _bb(closes: list[float], n: int = 20, k: float = 2.0) -> tuple[float, float, float, float]:
    """
    Bollinger Bands sur les n dernières closes.
    Retourne (upper, middle, lower, bbp) où :
      bbp = (close - lower) / (upper - lower)
      bbp < 0  → prix sous la bande basse  (extrême oversold)
      0..0.5   → moitié basse de la range
      0.5..1   → moitié haute de la range
      bbp > 1  → prix au-dessus de la bande haute (extrême overbought)
    """
    window = closes[-n:]
    mid    = sum(window) / len(window)
    var    = sum((p - mid) ** 2 for p in window) / len(window)
    std    = var ** 0.5
    upper  = mid + k * std
    lower  = mid - k * std
    bbp    = (closes[-1] - lower) / (upper - lower) if upper != lower else 0.5
    return upper, mid, lower, round(bbp, 3)

def _avg_vol(candles: list[dict], n: int = 20) -> float:
    vols = [c["v"] for c in candles[-n:]]
    return sum(vols) / len(vols) if vols else 0.0

# ── Signal multi-timeframe ────────────────────────────────────────────────────
def get_signal(coin: str) -> dict | None:
    try:
        c15 = fetch_candles(coin, INTERVAL_M15, 50)
        c5  = fetch_candles(coin, INTERVAL_M5,  50)
        c1  = fetch_candles(coin, INTERVAL_M1,  30)
        if len(c15) < 25 or len(c5) < 20 or len(c1) < 16:
            return None

        # M15 biais (EMA20 — plus stable que EMA8 pour filtrer la tendance)
        ema20_15 = _ema([c["c"] for c in c15], 20)
        bias_up = c15[-1]["c"] > ema20_15[-1]
        bias_dn = c15[-1]["c"] < ema20_15[-1]

        # M5 confirmation (VWAP + EMA8 + ATR + Bollinger Bands)
        closes5 = [c["c"] for c in c5]
        ema8_5  = _ema(closes5, 8)
        vwap5   = _vwap(c5)
        atr5    = _atr(c5, 14)
        last5   = closes5[-1]
        move_pct = abs(last5 - closes5[-2]) / closes5[-2] if closes5[-2] else 0
        _, _, _, bbp5 = _bb(closes5, n=20, k=2.0)

        confirm_long  = last5 > ema8_5[-1] and last5 > vwap5
        confirm_short = last5 < ema8_5[-1] and last5 < vwap5
        # MR : prix à la BB basse/haute → souvent hors VWAP par définition.
        # On exige seulement EMA8 (rebond court-terme confirmé), pas VWAP.
        confirm_long_mr  = last5 > ema8_5[-1]
        confirm_short_mr = last5 < ema8_5[-1]

        avg_vol5 = _avg_vol(c5)
        vol_ok   = c5[-1]["v"] >= avg_vol5 * VOL_MULTIPLIER
        move_ok  = move_pct >= MIN_MOVE_PCT

        # BBP : prix dans la zone basse pour BUY, haute pour SELL
        # 0.35 = tiers bas de la bande  |  0.65 = tiers haut
        bb_long_ok  = bbp5 <= 0.35
        bb_short_ok = bbp5 >= 0.65

        # M5 RSI (stable — pour BBO breakout, évite les spikes M1)
        rsi5 = _rsi(closes5, 14)

        # M1 entrée (RSI14 + volume)
        closes1  = [c["c"] for c in c1]
        rsi1     = _rsi(closes1, 14)
        avg_vol1 = _avg_vol(c1)
        vol1_ok  = c1[-1]["v"] >= avg_vol1 * VOL_MULTIPLIER

        # Confirmation directionnelle M5 :
        # MR (mean-reversion) : 2 bougies suffisent — la consolidation génère des candles mixtes
        # MOM (momentum)      : 3 bougies requises — signal de force, pas de compromis
        bars2_bull = all(c5[-i]["c"] > c5[-i]["o"] for i in range(1, 3))
        bars2_bear = all(c5[-i]["c"] < c5[-i]["o"] for i in range(1, 3))
        bars3_bull = all(c5[-i]["c"] > c5[-i]["o"] for i in range(1, 4))
        bars3_bear = all(c5[-i]["c"] < c5[-i]["o"] for i in range(1, 4))

        signal    = None
        score     = 0.0
        mode      = ""
        vol_ratio = c5[-1]["v"] / avg_vol5 if avg_vol5 else 1.0

        base_mr_long = vol_ok and move_ok and vol1_ok and bars2_bull   # BUY MR  : 2 bougies
        base_mr_sht  = vol_ok and move_ok and vol1_ok and bars2_bear   # SELL MR : 2 bougies
        base_mom_sht = vol_ok and move_ok and vol1_ok and bars3_bear   # SELL MOM: 3 bougies

        # ── BUY mean-reversion : RSI en zone neutre, prix bas de la bande ──────
        if bias_up and confirm_long_mr and base_mr_long and 30 <= rsi1 <= 70 and bbp5 <= 0.40:
            signal    = "buy"
            mode      = "MR"
            rsi_score = (70 - rsi1) / 40
            bb_score  = (0.40 - bbp5) / 0.40
            score     = round(rsi_score * 0.25 + vol_ratio * 0.4 + bb_score * 0.15 + 0.2, 3)

        # ── SELL mean-reversion : RSI neutre, prix haut de la bande ─────────────
        elif bias_dn and confirm_short_mr and base_mr_sht and 35 <= rsi1 <= 65 and bb_short_ok:
            signal    = "sell"
            mode      = "MR"
            rsi_score = (rsi1 - 35) / 30
            bb_score  = (bbp5 - 0.65) / 0.35
            score     = round(rsi_score * 0.25 + vol_ratio * 0.4 + bb_score * 0.15 + 0.2, 3)

        # ── SELL momentum : crash/tendance forte — prix sous la bande basse ─────
        # RSI extrême (<= 35) + prix cassant sous la BB + toute la structure baissière
        elif bias_dn and confirm_short and base_mom_sht and rsi1 <= 35 and bbp5 <= 0.35:
            signal    = "sell"
            mode      = "MOM"
            rsi_score = max(0.0, (35 - rsi1) / 35)   # plus RSI bas = momentum fort
            bb_score  = max(0.0, (0.35 - bbp5) / 0.35)
            score     = round(vol_ratio * 0.5 + rsi_score * 0.25 + bb_score * 0.25, 3)

        # ── BUY breakout : cassure haussière au-dessus de la BB haute ────────────
        # Uniquement en régime BULL — prix > BB haute (BBP ≥ 0.75) + trend ↑ + volume fort
        # Utilise rsi5 (M5, 14 périodes) — stable pendant les pompes, contrairement à rsi1
        # qui monte à 90+ en quelques minutes M1 et bloque tous les signaux de breakout.
        elif (_current_regime == "BULL" and bias_up and confirm_long
              and vol_ok and move_ok and 45 <= rsi5 <= 85 and bbp5 >= 0.75):
            signal    = "buy"
            mode      = "BBO"
            rsi_score = max(0.0, (85 - rsi5) / 40)      # rsi5=45 → 1.0 | rsi5=85 → 0.0
            bb_score  = min(1.0, (bbp5 - 0.75) / 0.50)  # cassure plus forte = mieux
            score     = round(vol_ratio * 0.50 + rsi_score * 0.30 + bb_score * 0.20, 3)

        bars3_str = ("3×↑" if bars3_bull else ("3×↓" if bars3_bear else
                     ("2×↑" if bars2_bull else ("2×↓" if bars2_bear else "✗"))))
        details = (f"M15={'↑' if bias_up else '↓'} M5={'↑' if confirm_long else '↓'} "
                   f"RSI={rsi1:.1f} BBP={bbp5:.2f} vol={c5[-1]['v']:.2f}/{avg_vol5:.2f} {bars3_str}"
                   + (f" [{mode}]" if mode else ""))

        if signal:
            # Appliquer les pénalités apprises des trades perdants
            penalty = get_penalty(coin, signal, rsi1, bbp5, _current_regime, mode)
            if penalty > 0:
                old_score = score
                score = max(0.0, round(score - penalty, 3))
                details += f" ⚠pén−{penalty}"
                log_scan(coin, "M15/M5/M1", signal, score,
                         details + f" [score {old_score:.3f}→{score:.3f} après leçon]")
            else:
                log_scan(coin, "M15/M5/M1", signal, score, details)
            return {
                "symbol": coin, "signal": signal, "score": score,
                "atr_m5": atr5, "price": last5,
                "conditions": {
                    "rsi": round(rsi1, 1), "bbp": round(bbp5, 3),
                    "vol_ratio": round(vol_ratio, 2),
                    "regime": _current_regime, "mode": mode,
                }
            }

        log_scan(coin, "M15/M5/M1", signal, score, details)
        return None

    except Exception as e:
        log_error(f"Signal {coin}: {e}")
        return None

# ── Balance (affichage uniquement) ───────────────────────────────────────────
def get_equity() -> float:
    user_state = info.user_state(WALLET_ADDRESS)
    cross        = user_state.get("crossMarginSummary") or user_state.get("marginSummary") or {}
    withdrawable = float(user_state.get("withdrawable") or 0)
    margin_used  = float(cross.get("totalMarginUsed") or 0)
    perp_total   = withdrawable + margin_used   # solde perp (libre + engagé)

    # Mode Unified Hyperliquid : le USDC Spot sert aussi de marge perp.
    # Il n'est PAS dans withdrawable — on l'ajoute séparément.
    spot_usdc = 0.0
    try:
        spot_state = info.spot_user_state(WALLET_ADDRESS)
        for b in spot_state.get("balances", []):
            if b["coin"] == "USDC":
                spot_usdc = float(b.get("total") or 0)
                break
    except Exception:
        pass

    total = perp_total + spot_usdc
    if total > 0:
        return total
    # Fallback pour comptes non-Unified : accountValue seul
    acv = float(cross.get("accountValue") or 0)
    return acv if acv > 0 else 0.0

def get_unrealized_pnl() -> float:
    try:
        positions = info.user_state(WALLET_ADDRESS).get("assetPositions", [])
        return sum(float(p["position"]["unrealizedPnl"]) for p in positions)
    except Exception:
        return 0.0

def get_market_regime() -> tuple[str, str, float, float]:
    """
    Analyse BTC et ETH sur 7 jours — 4 régimes (inspiré de tbot) :
      BULL      : BTC et ETH tous les deux > +5%  → trade normal, favorise BUY
      BEAR      : BTC et ETH tous les deux < -5%  → aucun nouveau trade
      RANGING   : les deux dans ±5%               → exige score élevé
      DIVERGING : l'un monte, l'autre descend      → aucun nouveau trade (marché incohérent)

    Retourne (regime, bias, btc_chg_pct, eth_chg_pct)
    """
    try:
        btc = fetch_candles("BTC", INTERVAL_D1, 8)
        eth = fetch_candles("ETH", INTERVAL_D1, 8)
        btc_chg = (btc[-1]["c"] - btc[0]["c"]) / btc[0]["c"] * 100
        eth_chg = (eth[-1]["c"] - eth[0]["c"]) / eth[0]["c"] * 100

        # DIVERGING : BTC et ETH se contredisent (>3% d'écart et dans des sens opposés)
        if btc_chg > 3 and eth_chg < -3:
            return "DIVERGING", "neutral", btc_chg, eth_chg
        if btc_chg < -3 and eth_chg > 3:
            return "DIVERGING", "neutral", btc_chg, eth_chg

        avg_chg = (btc_chg + eth_chg) / 2

        if avg_chg > 5:
            return "BULL",    "up",      btc_chg, eth_chg
        if avg_chg < -5:
            return "BEAR",    "down",    btc_chg, eth_chg
        return "RANGING",     "neutral", btc_chg, eth_chg
    except Exception:
        return "RANGING", "neutral", 0.0, 0.0   # safe default

def load_initial_balance(current_balance: float) -> float:
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE) as f:
                state = json.load(f)
            if "initial_balance" in state:
                return float(state["initial_balance"])
        except Exception:
            pass
    _save_initial_balance(current_balance)
    return current_balance

def _save_initial_balance(balance: float):
    state = {}
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE) as f:
                state = json.load(f)
        except Exception:
            pass
    state["initial_balance"] = round(balance, 4)
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)

def save_positions_state(positions: list):
    """Persiste les positions ouvertes dans bot_state.json après chaque changement."""
    state = {}
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE) as f:
                state = json.load(f)
        except Exception:
            pass
    state["open_positions"] = positions
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)

def load_positions_state() -> dict[str, dict]:
    """Charge l'état enrichi des positions depuis bot_state.json.
    Retourne {coin: position_dict} pour enrichir la reprise depuis l'API."""
    if not os.path.exists(STATE_FILE):
        return {}
    try:
        with open(STATE_FILE) as f:
            state = json.load(f)
        saved = state.get("open_positions", [])
        return {p["symbol"]: p for p in saved if "symbol" in p}
    except Exception:
        return {}

# ── Scan parallèle ────────────────────────────────────────────────────────────
def scan_all(open_coins: set[str]) -> list[dict]:
    candidates = [c for c in WATCHLIST if c not in open_coins]
    results = []
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(get_signal, c): c for c in candidates}
        for fut in as_completed(futs):
            sig = fut.result()
            if sig:
                results.append(sig)
    return results

def best_signal(sigs: list[dict]) -> dict | None:
    return max(sigs, key=lambda s: s["score"]) if sigs else None

# ── Taille de position ─────────────────────────────────────────────────────────
def calc_qty(coin: str, price: float) -> float:
    f = _flt(coin)
    notional = CAPITAL_PER_TRADE * LEVERAGE
    sz = notional / price
    factor = 10 ** f["sz_dec"]
    sz = math.floor(sz * factor) / factor
    return max(f["sz_step"], sz)

def round_px(price: float) -> float:
    if price >= 1000: return round(price, 1)
    if price >= 100:  return round(price, 2)
    if price >= 10:   return round(price, 3)
    if price >= 1:    return round(price, 4)
    return round(price, 6)

# ── Levier ────────────────────────────────────────────────────────────────────
def set_leverage(coin: str):
    try:
        res = exchange.update_leverage(LEVERAGE, coin, is_cross=True)
        if res.get("status") != "ok":
            print(f"  {C.BYLW}[WARN]{C.RST} Levier {coin}: {res}")
    except Exception as e:
        print(f"  {C.BYLW}[WARN]{C.RST} Levier {coin}: {e}")

# ── Alertes Discord ───────────────────────────────────────────────────────────
def send_discord_alert(title: str, description: str, color: int = 0xe74c3c):
    """Envoie une alerte embed sur Discord via webhook. Silencieux si pas configuré."""
    if not DISCORD_WEBHOOK_URL:
        return
    payload = {
        "embeds": [{
            "title": title,
            "description": description,
            "color": color,
            "footer": {"text": f"BotShortTrade · {WALLET_ADDRESS[:6]}…{WALLET_ADDRESS[-4:]}"},
            "timestamp": datetime.now(timezone.utc).isoformat()
        }]
    }
    try:
        r = requests.post(DISCORD_WEBHOOK_URL, json=payload, timeout=5)
        r.raise_for_status()
    except Exception:
        pass  # Ne jamais crasher le bot à cause de Discord


# ── Passer un ordre ───────────────────────────────────────────────────────────
def place_order(sig: dict, wb: Workbook) -> dict:
    coin   = sig["symbol"]
    atr    = sig["atr_m5"]
    is_buy = sig["signal"] == "buy"
    side   = "BUY" if is_buy else "SELL"

    set_leverage(coin)

    all_mids = info.all_mids()
    price = float(all_mids.get(coin, 0))
    if not price:
        raise RuntimeError(f"Prix indisponible: {coin}")

    qty = calc_qty(coin, price)

    if is_buy:
        sl = round_px(price - ATR_SL_MULT * atr)
        tp = round_px(price + ATR_TP_MULT * atr)
    else:
        sl = round_px(price + ATR_SL_MULT * atr)
        tp = round_px(price - ATR_TP_MULT * atr)

    _sc = C.BGRN if is_buy else C.BRED
    print(f"  {_sc}[{side}]{C.RST} {C.BOLD}{coin}{C.RST} {qty} @ ${price:.6g}  {C.RED}SL=${sl:.6g}{C.RST}  {C.GRN}TP=${tp:.6g}{C.RST}  {LEVERAGE}x")

    # Ouverture de la position au marché
    result = exchange.market_open(coin, is_buy, qty, px=None, slippage=0.02)
    if result.get("status") != "ok":
        raise RuntimeError(f"Ordre refusé {coin}: {result}")

    statuses  = result["response"]["data"]["statuses"]
    fill      = statuses[0].get("filled") or statuses[0].get("resting") or {}
    order_id  = str(fill.get("oid", f"hl-{coin}-{int(time.time())}"))
    entry_px  = float(fill.get("avgPx") or price)
    open_fee  = round(float(fill.get("fee", 0)), 4)

    print(f"  {C.BGRN}[OK]{C.RST} oid={order_id}  entrée=${entry_px:.6g}  frais=${open_fee}")

    # Notification Discord — ouverture de trade
    _ds = "🟢 LONG" if is_buy else "🔴 SHORT"
    send_discord_alert(
        f"{_ds} {coin} ouvert",
        f"Entrée: **${entry_px:.6g}** | SL: **${sl:.6g}** | TP: **${tp:.6g}**\n"
        f"Qty: {qty} | Notionnel: **${notional}** | {LEVERAGE}x levier",
        color=0x2ecc71 if is_buy else 0xe74c3c,
    )

    # Ordres SL / TP (trigger reduce-only)
    close_buy = not is_buy
    sl_lim = round_px(sl * 0.90) if is_buy else round_px(sl * 1.10)
    tp_lim = round_px(tp * 1.10) if is_buy else round_px(tp * 0.90)
    sl_oid = None

    for label, trig_px, lim_px, tpsl in [
        ("SL", sl, sl_lim, "sl"),
        ("TP", tp, tp_lim, "tp"),
    ]:
        try:
            r = exchange.order(
                coin, close_buy, qty,
                limit_px=lim_px,
                order_type={"trigger": {"triggerPx": float(trig_px), "isMarket": True, "tpsl": tpsl}},
                reduce_only=True,
            )
            if r.get("status") == "ok":
                _lc = C.RED if label == "SL" else C.GRN
                print(f"  {_lc}[{label}]{C.RST} ${trig_px:.6g} → OK")
                if label == "SL":
                    st = r["response"]["data"]["statuses"]
                    sl_oid = (st[0].get("resting") or st[0].get("filled") or {}).get("oid")
            else:
                print(f"  {C.BRED}[{label}] REJETÉ{C.RST} : {r}")
        except Exception as e:
            print(f"  {C.BRED}[{label}] ERREUR{C.RST} : {e}")

    notional = round(qty * entry_px, 2)
    pnl_tp   = round(abs(tp - entry_px) * qty, 2)
    pnl_sl   = round(-abs(sl - entry_px) * qty, 2)

    log_order(wb, order_id, coin, side, qty, entry_px, sl, tp, atr, notional, pnl_tp, pnl_sl)

    return {"id": order_id, "symbol": coin, "side": side,
            "entry": entry_px, "qty": qty, "sl": sl, "tp": tp,
            "sl_oid": sl_oid, "atr": atr, "open_fee": open_fee,
            "conditions": sig.get("conditions", {}),
            "opened_at": time.time()}

# ── Trailing stop ────────────────────────────────────────────────────────────
def trail_sl(position: dict, current_price: float):
    coin   = position["symbol"]
    is_buy = position["side"] == "BUY"
    entry  = position["entry"]
    atr    = position.get("atr", entry * 0.002)
    sl     = position["sl"]
    sl_oid = position.get("sl_oid")

    profit_atr = ((current_price - entry) / atr) if is_buy else ((entry - current_price) / atr)
    if profit_atr < TRAIL_TRIGGER:
        return

    new_sl = round_px(current_price - TRAIL_DIST * atr) if is_buy \
             else round_px(current_price + TRAIL_DIST * atr)

    # N'avancer le SL que dans le sens favorable
    if (is_buy and new_sl <= sl) or (not is_buy and new_sl >= sl):
        return

    # Annuler l'ancien SL
    if sl_oid:
        try:
            exchange.cancel(coin, int(sl_oid))
        except Exception as e:
            print(f"  {C.BYLW}[TRAIL]{C.RST} annulation SL : {e}")

    # Placer le nouveau SL
    close_buy = not is_buy
    sl_lim = round_px(new_sl * 0.90) if is_buy else round_px(new_sl * 1.10)
    try:
        r = exchange.order(
            coin, close_buy, position["qty"],
            limit_px=sl_lim,
            order_type={"trigger": {"triggerPx": float(new_sl), "isMarket": True, "tpsl": "sl"}},
            reduce_only=True,
        )
        if r.get("status") == "ok":
            st = r["response"]["data"]["statuses"]
            position["sl_oid"] = (st[0].get("resting") or {}).get("oid")
            position["sl"] = new_sl
            arrow = "↑" if is_buy else "↓"
            print(f"  {C.BCYN}[TRAIL {arrow}]{C.RST} {coin} SL={new_sl:.6g} (profit={profit_atr:.1f}×ATR)")
        else:
            print(f"  {C.BYLW}[TRAIL]{C.RST} rejeté : {r}")
    except Exception as e:
        print(f"  {C.BYLW}[TRAIL]{C.RST} erreur : {e}")

# ── Suivi de position ─────────────────────────────────────────────────────────
def check_position(wb: Workbook, position: dict) -> str | None:
    coin = position["symbol"]
    try:
        user_state = info.user_state(WALLET_ADDRESS)
        asset_pos  = user_state.get("assetPositions", [])
        pos = next(
            (p["position"] for p in asset_pos
             if p["position"]["coin"] == coin and abs(float(p["position"]["szi"])) > 0),
            None
        )

        if pos:
            mids  = info.all_mids()
            price = float(mids.get(coin, position["entry"]))
            pnl   = float(pos.get("unrealizedPnl", 0))
            pct   = (price - position["entry"]) / position["entry"] * 100
            if position["side"] == "SELL":
                pct = -pct
            _pc = _cp(pnl)
            print(f"  {C.DIM}[SUIVI]{C.RST} {C.BOLD}{coin}{C.RST} ${price:.6g} ({pct:+.2f}%)  P&L={_pc}${pnl:.2f}{C.RST}")
            update_order_live(wb, position["id"], price)
            trail_sl(position, price)
            return None

        # Position fermée — chercher le P&L dans les fills
        pnl, result = 0.0, "Fermé"
        try:
            fills   = info.user_fills(WALLET_ADDRESS)
            closing = [f for f in fills
                       if f["coin"] == coin and "Close" in f.get("dir", "")]
            if closing:
                recent = max(closing, key=lambda f: f["time"])
                close_fee = float(recent.get("fee", 0))
                open_fee  = position.get("open_fee", close_fee)
                gross     = float(recent.get("closedPnl", 0))
                pnl       = round(gross - open_fee - close_fee, 2)
                result    = "TP touché" if pnl >= 0 else "SL touché"
                print(f"  {C.DIM}[FRAIS] ouv=${open_fee:.4f}  ferm=${close_fee:.4f}  net=${pnl:.2f}{C.RST}")
        except Exception:
            pass

        _rc = C.BGRN if "TP" in result else C.BRED
        print(f"  {_rc}[{result}]{C.RST} {C.BOLD}{coin}{C.RST}  P&L={_cp(pnl)}${pnl:.2f}{C.RST}")

        # Notification Discord — fermeture de trade
        _emoji = "✅" if pnl >= 0 else "🛑"
        _pnl_s = f"+${pnl:.2f}" if pnl >= 0 else f"-${abs(pnl):.2f}"
        send_discord_alert(
            f"{_emoji} {coin} — {result}",
            f"P&L: **{_pnl_s}** | Entrée: **${position.get('entry', 0):.6g}**",
            color=0x2ecc71 if pnl >= 0 else 0xe74c3c,
        )

        record_lesson(position, result, pnl)   # apprendre des succès et des erreurs
        update_order_status(wb, position["id"], result, pnl)
        return result

    except Exception as e:
        log_error(f"check_position {coin}: {e}")
        return None

# ── Récupération positions à la reprise ──────────────────────────────────────
def recover_open_positions() -> tuple[list, float]:
    # Charger l'état enrichi du dernier run (sl_oid, atr, sl, tp, open_fee)
    saved_state = load_positions_state()

    try:
        user_state = info.user_state(WALLET_ADDRESS)
        asset_pos  = user_state.get("assetPositions", [])
    except Exception as e:
        log_error(f"recover: {e}")
        return [], 0.0

    open_pos = [p["position"] for p in asset_pos
                if p["position"]["coin"] in set(WATCHLIST)
                and abs(float(p["position"]["szi"])) > 0]
    if not open_pos:
        return [], 0.0

    recovered, capital_used = [], 0.0
    for p in open_pos:
        coin  = p["coin"]
        size  = float(p["szi"])
        side  = "BUY" if size > 0 else "SELL"
        entry = float(p.get("entryPx") or 0)
        qty   = abs(size)

        # Tenter d'enrichir avec l'état sauvegardé (sl_oid, atr réel, sl/tp précis)
        saved = saved_state.get(coin, {})
        enriched = saved and saved.get("side") == side
        if enriched:
            sl       = saved.get("sl",       (entry - ATR_SL_MULT * entry * 0.002) if side == "BUY" else (entry + ATR_SL_MULT * entry * 0.002))
            tp       = saved.get("tp",       (entry + ATR_TP_MULT * entry * 0.002) if side == "BUY" else (entry - ATR_TP_MULT * entry * 0.002))
            atr      = saved.get("atr",      entry * 0.002)
            sl_oid   = saved.get("sl_oid")
            open_fee = saved.get("open_fee", 0.0)
            oid      = saved.get("id",       f"recovered-{coin}")
        else:
            atr_fb   = entry * 0.002
            sl       = (entry - ATR_SL_MULT * atr_fb) if side == "BUY" else (entry + ATR_SL_MULT * atr_fb)
            tp       = (entry + ATR_TP_MULT * atr_fb) if side == "BUY" else (entry - ATR_TP_MULT * atr_fb)
            atr      = atr_fb
            sl_oid   = None
            open_fee = 0.0
            oid      = f"recovered-{coin}"

        position = {"id": oid, "symbol": coin, "side": side,
                    "entry": entry, "qty": qty, "sl": sl, "tp": tp,
                    "sl_oid": sl_oid, "atr": atr, "open_fee": open_fee}
        recovered.append(position)
        capital_used += CAPITAL_PER_TRADE

        _sc  = C.BGRN if side == "BUY" else C.BRED
        _enr = (C.BGRN + "état complet" + C.RST) if enriched else (C.BYLW + "état partiel" + C.RST)
        print(f"  {C.BCYN}[REPRISE]{C.RST} {coin} {_sc}{side}{C.RST} @ ${entry:.6g}  ({_enr})")

    return recovered, capital_used

def reconcile_zombie_orders(wb: Workbook, open_positions: list) -> None:
    ws = wb["Trades"]
    open_ids = {str(p["id"]) for p in open_positions}
    for row in ws.iter_rows(min_row=2, values_only=False):
        if str(row[13].value) == "Ordre placé" and str(row[0].value) not in open_ids:
            row[13].value = "Fermé (inconnu)"
            for cell in row:
                cell.fill = _FILLS["sl"]
    _safe_save(wb)

# ── Affichage scan ────────────────────────────────────────────────────────────
# ── P&L portefeuille ──────────────────────────────────────────────────────────
def calc_portfolio_pnl(wb: Workbook) -> tuple[float, float]:
    ws = wb["Trades"]
    now = datetime.now(timezone.utc)
    total, weekly = 0.0, 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        pnl = row[14]   # colonne O — P&L final $
        if pnl is None or pnl == "":
            continue
        try:
            pnl = float(pnl)
        except (ValueError, TypeError):
            continue
        total += pnl
        date_str = str(row[1] or "")   # colonne B — Date/Heure
        try:
            trade_dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
            if (now - trade_dt).days < 7:
                weekly += pnl
        except ValueError:
            pass
    return round(total, 2), round(weekly, 2)

def _print_scan_summary():
    global _last_scan
    if not _scan_log:
        return
    print(f"\n  {C.DIM}── Scan ({len(_scan_log)} symboles) ──{C.RST}")
    for e in _scan_log:
        sig = e["signal"] or "–"
        _sc = C.BGRN if sig == "buy" else (C.BRED if sig == "sell" else C.DIM)
        print(f"    {C.BOLD}{e['coin']:5s}{C.RST} {_sc}{sig:4s}{C.RST}  score={e['score']:.3f}  {C.DIM}{e['details']}{C.RST}")
    _last_scan = list(_scan_log)   # snapshot pour le dashboard
    _scan_log.clear()

# ── Génération du dashboard HTML ──────────────────────────────────────────────
def write_dashboard(wb: Workbook, open_positions: list, balance: float,
                    initial_balance: float, pnl_total: float, pnl_week: float,
                    unrealized: float, avail_capital: float,
                    regime: str, btc_chg: float, eth_chg: float):
    try:
        ws = wb["Trades"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[14] is not None and r[14] != ""]
        wins   = [r for r in rows if float(r[14] or 0) > 0]
        losses = [r for r in rows if float(r[14] or 0) < 0]
        total_trades = len(rows)
        win_rate     = len(wins) / total_trades * 100 if total_trades else 0.0
        gross_profit = sum(float(r[14]) for r in wins)
        gross_loss   = abs(sum(float(r[14]) for r in losses))
        profit_factor = round(gross_profit / gross_loss, 2) if gross_loss > 0 else 0.0
        net_pnl  = pnl_total + unrealized
        net_pct  = net_pnl / initial_balance * 100 if initial_balance > 0 else 0.0
        now_str  = datetime.now(timezone.utc).strftime("%d/%m/%Y à %H:%M:%S UTC")

        def _pct_col(v): return "green" if v >= 0 else "red"
        def _val_col(v): return "green" if v > 0 else ("red" if v < 0 else "white")
        def _fmt(v):     return f"+${v:.2f}" if v > 0 else f"-${abs(v):.2f}"

        regime_color = {"BULL": "#4ade80", "BEAR": "#f87171", "DIVERGING": "#f87171", "RANGING": "#fbbf24"}.get(regime, "#e0e0e0")

        # ── Positions ouvertes ─────────────────────────────────────────────────
        pos_rows = ""
        for p in open_positions:
            side_badge = f'<span class="buy-badge">BUY</span>' if p["side"] == "BUY" else f'<span class="sell-badge">SELL</span>'
            entry = p.get("entry", 0)
            sl    = p.get("sl", 0)
            tp    = p.get("tp", 0)
            pos_rows += f"""
            <tr>
              <td><span class="coin-badge">{p['symbol']}/USD</span></td>
              <td>{side_badge}</td>
              <td>${entry:.4f}</td>
              <td>${sl:.4f}</td>
              <td>${tp:.4f}</td>
              <td><span class="pending-badge">Ouvert</span></td>
            </tr>"""
        if not pos_rows:
            pos_rows = '<tr><td colspan="6" style="text-align:center;color:#555">Aucune position ouverte</td></tr>'

        # ── Dernier scan ───────────────────────────────────────────────────────
        scan_rows = ""
        for e in _last_scan:
            sig = e.get("signal") or "–"
            if sig == "buy":
                sig_badge = '<span class="buy-badge">BUY</span>'
            elif sig == "sell":
                sig_badge = '<span class="sell-badge">SELL</span>'
            else:
                sig_badge = '<span class="flat-badge">FLAT</span>'
            score = e.get("score", 0.0)
            details = e.get("details", "")
            ts = e.get("ts", "")
            scan_rows += f"""
            <tr>
              <td><span class="coin-badge">{e['coin']}/USD</span></td>
              <td>{sig_badge}</td>
              <td style="font-size:0.85em;color:#aaa">{score:.3f}</td>
              <td style="font-size:0.82em;color:#666">{details}</td>
              <td style="font-size:0.82em;color:#555">{ts}</td>
            </tr>"""
        if not scan_rows:
            scan_rows = '<tr><td colspan="5" style="text-align:center;color:#555">En attente du prochain scan…</td></tr>'

        # ── Historique trades ──────────────────────────────────────────────────
        all_rows_sorted = list(reversed(rows))[:40]
        cumul = pnl_total
        trade_rows = ""
        running_total = 0.0
        for r in reversed(all_rows_sorted):
            running_total += float(r[14] or 0)
        cumul_run = running_total
        for r in all_rows_sorted:
            pnl_val  = float(r[14] or 0)
            cumul_run -= pnl_val   # we go reverse so subtract
            pnl_pct  = pnl_val / initial_balance * 100 if initial_balance > 0 else 0
            cumul_display = cumul_run + pnl_val   # cumul at this trade
            pnl_cls  = "gain" if pnl_val > 0 else "loss"
            cum_cls  = "gain" if cumul_display >= 0 else "loss"
            side_badge = f'<span class="buy-badge">BUY</span>' if str(r[3]).upper() == "BUY" else f'<span class="sell-badge">SELL</span>'
            status = str(r[13] or "")
            if "TP" in status.upper():
                st_badge = '<span class="tp-badge">TP ✓</span>'
            elif "SL" in status.upper():
                st_badge = '<span class="sl-badge">SL ✗</span>'
            elif "TRAIL" in status.upper():
                st_badge = '<span class="tp-badge">Trail ✓</span>'
            else:
                st_badge = f'<span class="pending-badge">{status}</span>'
            entry_px = r[5] or 0
            trade_rows += f"""
            <tr>
              <td style="font-size:0.82em;color:#777">{r[1] or ""}</td>
              <td><span class="coin-badge">{r[2] or ""}</span></td>
              <td>{side_badge}</td>
              <td>${float(entry_px):.4f}</td>
              <td class="{pnl_cls}"><strong>{_fmt(pnl_val)}</strong> <span style="opacity:.7;font-size:.82em">({pnl_pct:+.2f}%)</span></td>
              <td>{st_badge}</td>
            </tr>"""
        if not trade_rows:
            trade_rows = '<tr><td colspan="6" style="text-align:center;color:#555">Aucun trade fermé encore</td></tr>'

        # ── Leçons apprises ────────────────────────────────────────────────────
        MIN_SAMPLES = 5
        les_data = _load_lessons().get("buckets", {})
        all_buckets = sorted(les_data.items(),
                             key=lambda x: x[1]["w"] / max(1, x[1]["w"] + x[1]["l"]))
        lessons_rows = ""
        for k, v in all_buckets[:20]:
            total = v["w"] + v["l"]
            wr    = v["w"] / total if total else 0
            pen   = round(min((0.40 - wr) * 0.5, 0.25), 3) if wr < 0.40 and total >= MIN_SAMPLES else 0.0
            wr_color = "#f87171" if wr < 0.35 else ("#fbbf24" if wr < 0.50 else "#4ade80")
            pen_str  = f'<span style="color:#f87171;font-weight:600">−{pen}</span>' if pen > 0 else '<span style="color:#555">0 (apprentissage…)</span>'
            pnl_v    = v["pnl"]
            pnl_cls  = "gain" if pnl_v > 0 else "loss"
            lessons_rows += f"""
            <tr>
              <td style="font-family:monospace;font-size:0.85em;color:#ccc">{k}</td>
              <td style="font-weight:700;color:{wr_color}">{wr*100:.0f}%</td>
              <td style="color:#888">{v['w']}W / {v['l']}L ({total})</td>
              <td class="{pnl_cls}">{_fmt(pnl_v)}</td>
              <td>{pen_str}</td>
            </tr>"""
        if not lessons_rows:
            lessons_rows = '<tr><td colspan="5" style="text-align:center;color:#555">Aucune leçon encore — les données s\'accumulent après les premiers trades</td></tr>'

        html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta http-equiv="refresh" content="15">
  <title>Bot Crypto — Dashboard</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Segoe UI', sans-serif; background: #0f1117; color: #e0e0e0; padding: 24px; }}
    h1 {{ font-size: 1.6em; color: #fff; margin-bottom: 4px; }}
    .subtitle {{ color: #888; font-size: 0.9em; margin-bottom: 28px; }}
    .cards {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 32px; }}
    .card {{ background: #1a1d27; border-radius: 12px; padding: 20px 24px; flex: 1; min-width: 140px; border: 1px solid #2a2d3a; }}
    .card .label {{ font-size: 0.78em; color: #888; text-transform: uppercase; letter-spacing: .05em; }}
    .card .value {{ font-size: 2em; font-weight: 700; margin-top: 6px; }}
    .card .sub {{ font-size: 0.82em; color: #888; margin-top: 3px; }}
    .value.green {{ color: #4ade80; }} .value.red {{ color: #f87171; }}
    .value.white {{ color: #fff; }}  .value.yellow {{ color: #fbbf24; }}
    .section {{ margin-bottom: 32px; }}
    .section h2 {{ font-size: 1.05em; color: #aaa; margin-bottom: 12px; border-bottom: 1px solid #2a2d3a; padding-bottom: 8px; }}
    .regime-bar {{ display:inline-flex; align-items:center; gap:10px; background:#1a1d27; border-radius:8px; padding:10px 18px; border:1px solid #2a2d3a; margin-bottom:20px; }}
    .regime-dot {{ width:10px; height:10px; border-radius:50%; background:{regime_color}; }}
    .regime-name {{ font-weight:700; color:{regime_color}; font-size:1.05em; }}
    table {{ width: 100%; border-collapse: collapse; background: #1a1d27; border-radius: 10px; overflow: hidden; }}
    th {{ background: #12151f; color: #888; font-size: 0.78em; text-transform: uppercase; padding: 10px 14px; text-align: left; letter-spacing: .05em; }}
    td {{ padding: 10px 14px; border-top: 1px solid #2a2d3a; font-size: 0.9em; }}
    tr:hover td {{ background: #20243a; }}
    .gain {{ color: #4ade80; font-weight: 600; }} .loss {{ color: #f87171; font-weight: 600; }}
    .coin-badge    {{ background:#2a2d3a; color:#e0e0e0; padding:2px 8px; border-radius:20px; font-size:0.82em; font-weight:600; }}
    .buy-badge     {{ background:#14532d; color:#4ade80; padding:2px 10px; border-radius:20px; font-size:0.82em; font-weight:700; }}
    .sell-badge    {{ background:#450a0a; color:#f87171; padding:2px 10px; border-radius:20px; font-size:0.82em; font-weight:700; }}
    .flat-badge    {{ background:#292524; color:#a8a29e; padding:2px 10px; border-radius:20px; font-size:0.82em; }}
    .tp-badge      {{ background:#14532d; color:#4ade80; padding:2px 10px; border-radius:20px; font-size:0.82em; }}
    .sl-badge      {{ background:#450a0a; color:#f87171; padding:2px 10px; border-radius:20px; font-size:0.82em; }}
    .pending-badge {{ background:#1c1917; color:#fbbf24; padding:2px 10px; border-radius:20px; font-size:0.82em; }}
    .footer {{ color: #444; font-size: 0.78em; margin-top: 24px; text-align: center; }}
    /* ── Tooltips ─────────────────────────────────────────── */
    [data-tip]          {{ cursor: help; }}
    .label [data-tip]   {{ border-bottom: 1px dotted #555; padding-bottom: 1px; }}
    th[data-tip]        {{ cursor: help; }}
    #_tip {{
      position: fixed;
      background: #161b2e;
      color: #d4daf0;
      font-size: 1em;
      line-height: 1.65;
      padding: 11px 15px;
      border-radius: 9px;
      border: 1px solid #2e3350;
      max-width: 320px;
      z-index: 9999;
      box-shadow: 0 6px 28px rgba(0,0,0,.8);
      pointer-events: none;
      display: none;
      white-space: normal;
      text-transform: none;
      letter-spacing: 0;
      font-weight: 400;
    }}
  </style>
</head>
<body>
  <h1>📊 Bot Crypto — Dashboard</h1>
  <p class="subtitle">Mis à jour le {now_str} · Rafraîchissement auto toutes les 15s · {'⚠️ MAINNET (argent réel)' if not HL_TESTNET else '🟡 TESTNET (simulation)'}</p>

  <div class="cards">
    <div class="card">
      <div class="label"><span data-tip="Valeur totale du compte (perp + USDC spot). Sert de base pour calculer la taille des positions. Réf = capital initial au démarrage du bot.">Équité</span></div>
      <div class="value white">${balance:.2f}</div>
      <div class="sub">Réf: ${initial_balance:.2f}</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Capital libre pour ouvrir de nouvelles positions = Équité − capital engagé dans les positions ouvertes.">Dispo</span></div>
      <div class="value {'red' if avail_capital <= 0 else 'yellow' if avail_capital < CAPITAL_PER_TRADE * 2 else 'white'}">${avail_capital:.2f}</div>
      <div class="sub">{len(open_positions)} position(s) ouverte(s)</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Profit/Perte net depuis le démarrage : gains réalisés (trades fermés) + non-réalisés (positions en cours). Exprimé en % du capital initial.">P&amp;L Net</span></div>
      <div class="value {_val_col(net_pnl)}">{_fmt(net_pnl)}</div>
      <div class="sub">{net_pct:+.2f}% du capital initial</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Gains réalisés sur les 7 derniers jours (trades fermés uniquement, hors positions ouvertes).">P&amp;L 7 jours</span></div>
      <div class="value {_val_col(pnl_week)}">{_fmt(pnl_week)}</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Nombre total de positions clôturées (TP = Take Profit touché, SL = Stop Loss touché). W = gagnants, L = perdants.">Trades fermés</span></div>
      <div class="value white">{total_trades}</div>
      <div class="sub">{len(wins)}W / {len(losses)}L</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Pourcentage de trades fermés avec un profit. 84%+ = excellent. Le bot vise >70% grâce aux filtres RSI, BBP et volume.">Win Rate</span></div>
      <div class="value {'green' if win_rate >= 50 else 'red'}">{win_rate:.1f}%</div>
    </div>
    <div class="card">
      <div class="label"><span data-tip="Gains bruts / Pertes brutes. >2 = très rentable, >1 = rentable, <1 = perte nette. Idéalement >3 pour un scalper.">Profit Factor</span></div>
      <div class="value {'green' if profit_factor >= 1.5 else 'yellow' if profit_factor >= 1 else 'red'}">{profit_factor:.2f}</div>
    </div>
  </div>

  <div class="section">
    <h2>🌍 Régime de marché</h2>
    <div class="regime-bar">
      <div class="regime-dot"></div>
      <span class="regime-name" data-tip="BULL = BTC+ETH en hausse sur 7j → signaux BUY prioritaires (+ SELL MR score>0.75). BEAR = baissiers → SELL MOM. RANGING = lateral → Mean-Reversion. DIVERGING = BTC et ETH en directions opposées.">{regime}</span>
      <span style="color:#666;font-size:.9em">BTC {btc_chg:+.1f}% · ETH {eth_chg:+.1f}% (7j)</span>
    </div>
  </div>

  <div class="section">
    <h2>📂 Positions ouvertes</h2>
    <table>
      <thead>
        <tr><th>Crypto</th><th>Direction</th><th data-tip="Prix d'ouverture de la position.">Entrée</th><th data-tip="Niveau de prix qui déclenche la fermeture automatique pour limiter la perte. Calculé à 1.5×ATR sous/sur l'entrée.">Stop Loss</th><th data-tip="Prix cible pour clôturer avec profit. Calculé à 4×ATR (ratio gain/risque ≈ 2.7). Le prix doit l'atteindre pour clôturer.">Take Profit</th><th>Statut</th></tr>
      </thead>
      <tbody>{pos_rows}</tbody>
    </table>
  </div>

  <div class="section">
    <h2>🔍 Dernier scan du marché</h2>
    <table>
      <thead>
        <tr><th>Crypto</th><th data-tip="BUY = signal d'achat détecté · SELL = signal de vente · FLAT = aucun signal ce cycle">Signal</th><th data-tip="Force du signal de 0 à ~2. Plus c'est élevé, plus le bot est confiant dans l'entrée.">Score</th><th data-tip="M15/M5 = tendance haussière (↑) ou baissière (↓) sur 15min et 5min. RSI = momentum 14 périodes (>70 suracheté, <30 survendu). BBP = position dans les Bollinger Bands (0=bord bas, 1=bord haut, négatif=sous la bande). vol = volume actuel / volume moyen (✗=insuffisant, Nx↓=en déclin)">Détails</th><th>Heure</th></tr>
      </thead>
      <tbody>{scan_rows}</tbody>
    </table>
  </div>

  <div class="section">
    <h2>🧠 Leçons apprises (patterns pénalisés)</h2>
    <table>
      <thead>
        <tr><th data-tip="Catégorie de trade apprise par le bot. Ex: coin_AVAX|buy = achats AVAX · rsi70-75|buy|BBO = breakout avec RSI 70-75 · bbp&gt;1.0 = prix au-dessus de la Bollinger haute · regime_BULL = trade en régime BULL">Pattern</th><th data-tip="Taux de réussite du pattern. Rouge &lt;35% (dangereux), orange &lt;50% (faible), vert &ge;50% (bon).">Win Rate</th><th>W/L</th><th>P&amp;L cumulé</th><th data-tip="Malus sur le score si WR &lt;40% avec &ge;5 trades. Plus le malus est fort, moins le bot entre dans ce type de trade. Max -0.25.">Pénalité score</th></tr>
      </thead>
      <tbody>{lessons_rows}</tbody>
    </table>
  </div>

  <div class="section">
    <h2>📋 Historique des trades (40 derniers)</h2>
    <table>
      <thead>
        <tr><th>Date</th><th>Crypto</th><th>Direction</th><th>Prix entrée</th><th>P&amp;L</th><th>Statut</th></tr>
      </thead>
      <tbody>{trade_rows}</tbody>
    </table>
  </div>

  <p class="footer">BotShortTrade · Hyperliquid DEX · Wallet {WALLET_ADDRESS[:6]}…{WALLET_ADDRESS[-4:]}</p>
<script>
(function(){{
  var tip=document.createElement('div'); tip.id='_tip';
  document.body.appendChild(tip);
  function show(e){{
    var txt=this.getAttribute('data-tip');
    if(!txt)return;
    tip.textContent=txt;
    tip.style.display='block';
    move(e);
  }}
  function move(e){{
    var m=16, w=tip.offsetWidth, h=tip.offsetHeight;
    var vw=window.innerWidth, vh=window.innerHeight;
    var x=e.clientX+m, y=e.clientY-h-m;
    if(x+w>vw-10) x=e.clientX-w-m;
    if(y<10) y=e.clientY+m;
    if(x<10) x=10;
    tip.style.left=x+'px'; tip.style.top=y+'px';
  }}
  function hide(){{ tip.style.display='none'; }}
  document.querySelectorAll('[data-tip]').forEach(function(el){{
    el.addEventListener('mouseenter',show);
    el.addEventListener('mousemove',move);
    el.addEventListener('mouseleave',hide);
  }});
}})();
</script>
</body>
</html>"""

        with open("dashboard.html", "w", encoding="utf-8") as f:
            f.write(html)
    except Exception as e:
        log_error(f"write_dashboard: {e}")


# ── Boucle principale ─────────────────────────────────────────────────────────
def run():
    print(f"\n{C.BCYN}{'═' * 60}{C.RST}")
    print(f"  {C.BOLD}BotShortTrade  –  Hyperliquid DEX{C.RST}")
    print(f"{C.BCYN}{'═' * 60}{C.RST}")

    wb = init_excel()
    init_exchange_info()

    # ── Calibrage du capital sur le solde réel ────────────────────────────────
    global MAX_LIQUIDITY, CAPITAL_PER_TRADE
    try:
        balance = get_equity()
        print(f"  Balance : {C.BOLD}${balance:.2f} USDC{C.RST}")
    except Exception as e:
        log_error(f"Balance: {e}")
        balance = _MAX_LIQ_CFG if _MAX_LIQ_CFG > 0 else 100.0

    if _MAX_LIQ_CFG > 0:
        # Plafond manuel — utile pour limiter l'exposition si le portefeuille est grand
        MAX_LIQUIDITY = min(_MAX_LIQ_CFG, balance)
        print(f"  {C.BYLW}Capital plafonné à ${MAX_LIQUIDITY:.2f} (config) sur ${balance:.2f} disponibles{C.RST}")
    else:
        # Mode auto : gère tout le portefeuille
        MAX_LIQUIDITY = balance
        print(f"  {C.BGRN}Capital auto : ${MAX_LIQUIDITY:.2f} USDC (solde complet){C.RST}")

    CAPITAL_PER_TRADE = MAX_LIQUIDITY * CAPITAL_PCT
    print(f"  {C.DIM}Par trade : {CAPITAL_PCT*100:.0f}% × ${MAX_LIQUIDITY:.2f} = {C.RST}{C.BOLD}${CAPITAL_PER_TRADE:.2f}{C.RST}")

    initial_balance = load_initial_balance(balance)
    print(f"  {C.DIM}Solde initial (référence % P&L) : ${initial_balance:.2f} USDC{C.RST}")

    open_positions, capital_in_use = recover_open_positions()
    if open_positions:
        reconcile_zombie_orders(wb, open_positions)
        print(f"  {C.BCYN}Positions reprises : {len(open_positions)}{C.RST}")
    else:
        print(f"  {C.DIM}Aucune position ouverte à la reprise{C.RST}")

    LOOP_SECONDS = 60
    print(f"\n  {C.DIM}Démarrage — scan toutes les {LOOP_SECONDS}s{C.RST}\n")

    while True:
        loop_start = time.time()
        ts      = datetime.now(timezone.utc).strftime("%H:%M:%S")
        now_min = datetime.now(timezone.utc).minute   # défini tôt — utilisé avant la section régime
        print(f"\n{C.DIM}[{ts}] ──────────────────────────────────────────────────{C.RST}")

        # Suivi des positions ouvertes
        closed_this_loop = []
        for pos in list(open_positions):
            result = check_position(wb, pos)
            if result is not None:
                closed_this_loop.append(pos)

        for pos in closed_this_loop:
            open_positions.remove(pos)
            capital_in_use = max(0.0, capital_in_use - CAPITAL_PER_TRADE)

        # Persister l'état si des positions ont changé ce cycle
        if closed_this_loop:
            save_positions_state(open_positions)

        # Balance à jour
        try:
            balance = get_equity()
        except Exception:
            pass

        # ── Auto-recalibrage du capital ───────────────────────────────────────────
        # Recalibrer si le solde réel s'écarte significativement de MAX_LIQUIDITY
        # (baisse > 20% OU hausse > 20%) ET aucune position ouverte.
        # Hausse = dépôt ou profit accumulé. Baisse = perte ou retrait.
        balance_drift = abs(balance - MAX_LIQUIDITY) / max(MAX_LIQUIDITY, 0.01)
        if balance_drift > 0.20 and len(open_positions) == 0:
            old_max = MAX_LIQUIDITY
            MAX_LIQUIDITY      = max(balance, 0.01)
            CAPITAL_PER_TRADE  = MAX_LIQUIDITY * CAPITAL_PCT
            capital_in_use     = 0.0
            direction = "↑" if balance > old_max else "↓"
            print(f"  {C.BYLW}[RECALIBRAGE {direction}]{C.RST} Solde ${old_max:.2f} → ${MAX_LIQUIDITY:.2f} | "
                  f"par trade ${CAPITAL_PER_TRADE:.2f}")

        # Disponible = min(tracking interne, cash libre réel)
        # Le solde Hyperliquid (balance) inclut la marge des positions ouvertes.
        # Si balance < capital_in_use, le compte est en sous-marge → bloquer tout nouveau trade.
        internal_avail = MAX_LIQUIDITY - capital_in_use
        real_free      = max(0.0, balance - capital_in_use)
        avail_capital  = min(internal_avail, real_free)

        # Alerte divergence — seulement toutes les 5 min pour ne pas spammer
        if (internal_avail > 0 and real_free < internal_avail * 0.8
                and now_min % 5 == 0):
            print(f"  {C.BRED}[ALERTE]{C.RST} Dispo interne ${internal_avail:.2f} vs "
                  f"cash libre ${real_free:.2f} — {len(open_positions)} position(s) en marge")

        max_slots     = MAX_TRADES if MAX_TRADES > 0 else 999
        slots_left    = max_slots - len(open_positions)

        pnl_total, pnl_week = calc_portfolio_pnl(wb)
        unrealized = get_unrealized_pnl()
        net_pnl = pnl_total + unrealized
        net_pct = (net_pnl / initial_balance * 100) if initial_balance > 0 else 0.0
        _dispo_c = C.BRED if avail_capital <= 0 else (C.BYLW if avail_capital < CAPITAL_PER_TRADE * 2 else C.WHT)
        print(f"  Positions: {C.BOLD}{len(open_positions)}{C.RST}  Dispo: {_dispo_c}${avail_capital:.2f}{C.RST}  Équité: {C.BOLD}${balance:.2f}{C.RST}")
        print(f"  P&L réalisé: {_cp(pnl_total)}${pnl_total:+.2f}{C.RST}  |  Non réalisé: {_cp(unrealized)}${unrealized:+.2f}{C.RST}  |  Net: {_cp(net_pnl)}{C.BOLD}${net_pnl:+.2f} ({net_pct:+.1f}%){C.RST}")
        print(f"  P&L 7j: {_cp(pnl_week)}${pnl_week:+.2f}{C.RST}")

        # Filtre horaire (heures UTC)
        hour_utc = datetime.now(timezone.utc).hour
        in_session = SESSION_START_UTC <= hour_utc < SESSION_END_UTC

        # Régime de marché — mis à jour toutes les 30 min (candles journalières, pas besoin de plus)
        now_min = datetime.now(timezone.utc).minute
        if "_regime_cache" not in run.__dict__ or now_min % 30 == 0:
            run._regime_cache = get_market_regime()
        regime, bias, btc_chg, eth_chg = run._regime_cache
        global _current_regime
        _current_regime = regime   # partagé avec get_signal() pour les leçons
        _rc = (C.BGRN if regime == "BULL" else
               C.BRED if regime in ("BEAR", "DIVERGING") else C.BYLW)
        print(f"  Régime: {_rc}{regime}{C.RST}  BTC={btc_chg:+.1f}%  ETH={eth_chg:+.1f}%")

        # Limite de perte journalière
        _, pnl_today = calc_portfolio_pnl(wb)
        daily_loss_limit = -MAX_LIQUIDITY * MAX_DAILY_LOSS_PCT / 100
        daily_loss_hit = pnl_today <= daily_loss_limit
        if daily_loss_hit:
            print(f"  {C.BRED}[STOP]{C.RST} Perte journalière atteinte (${pnl_today:.2f} ≤ ${daily_loss_limit:.2f})")

        # Bloquer les nouveaux trades en régime dangereux
        regime_blocked = regime in ("BEAR", "DIVERGING")
        if regime_blocked:
            print(f"  {C.BRED}[RÉGIME]{C.RST} {regime} — aucun nouveau trade")

        # Raison de non-scan (lisible)
        if not in_session:
            print(f"  {C.BYLW}[PAUSE]{C.RST} Hors session ({hour_utc}h UTC, actif {SESSION_START_UTC}h-{SESSION_END_UTC}h)")

        # Avertissement solde insuffisant (< 2 trades possibles, aucune position ouverte)
        if balance < CAPITAL_PER_TRADE * 2 and len(open_positions) == 0 and now_min % 5 == 0:
            print(f"  {C.BRED}[SOLDE BAS]{C.RST} Équité ${balance:.2f} — "
                  f"insuffisant pour trader (min recommandé ${CAPITAL_PER_TRADE * 3:.2f}). "
                  f"Déposer des fonds pour reprendre.")

        # Chercher un nouveau signal
        if slots_left > 0 and avail_capital >= CAPITAL_PER_TRADE and in_session and not daily_loss_hit and not regime_blocked:
            open_coins = {p["symbol"] for p in open_positions}
            signals    = scan_all(open_coins)
            _print_scan_summary()

            # Filtrage par régime :
            # BULL     → BUY dans la tendance + SELL MR autorisé (suracheté, score ≥ 0.75)
            #            SELL MOM bloqué (ne pas shorter un crash en plein BULL)
            # RANGING  → exige score ≥ 0.75 (marché sans direction claire)
            min_score = 0.75 if regime == "RANGING" else 0.0
            if regime == "BULL":
                signals = [s for s in signals if
                           s["signal"] == "buy" or
                           (s["signal"] == "sell" and
                            s.get("conditions", {}).get("mode") == "MR" and
                            s["score"] >= 0.75)]
            signals = [s for s in signals if s["score"] >= min_score]

            # Ouvrir jusqu'à slots_left positions si plusieurs bons signaux disponibles
            opened_this_cycle = 0
            while signals and slots_left > 0 and avail_capital >= CAPITAL_PER_TRADE:
                sig = best_signal(signals)
                if not sig:
                    break
                signals = [s for s in signals if s["symbol"] != sig["symbol"]]
                _ss = C.BGRN if sig["signal"] == "buy" else C.BRED
                print(f"\n  {_ss}[SIGNAL] {sig['symbol']} {sig['signal'].upper()}{C.RST}  score={sig['score']:.3f}  régime={regime}")
                try:
                    pos = place_order(sig, wb)
                    open_positions.append(pos)
                    capital_in_use += CAPITAL_PER_TRADE
                    avail_capital  -= CAPITAL_PER_TRADE
                    slots_left     -= 1
                    opened_this_cycle += 1
                    save_positions_state(open_positions)
                except Exception as e:
                    log_error(f"place_order: {e}")
                    traceback.print_exc()
                    break  # ne pas tenter un 2e ordre si le 1er échoue

            if opened_this_cycle == 0:
                print(f"  {C.DIM}Pas de signal retenu ce cycle{C.RST}")
        elif slots_left <= 0:
            print(f"  {C.DIM}Toutes les positions sont ouvertes — scan ignoré{C.RST}")
            _scan_log.clear()
        elif avail_capital < CAPITAL_PER_TRADE:
            print(f"  {C.DIM}Capital insuffisant (${avail_capital:.2f} < ${CAPITAL_PER_TRADE:.2f}) — scan ignoré{C.RST}")
            _scan_log.clear()
        else:
            _scan_log.clear()

        # Afficher le résumé des leçons toutes les 10 min
        if now_min % 10 == 0:
            print_lessons_summary()

        # ── Détection d'anomalies + alertes Discord ───────────────────────────────
        # 1) Flash crash : équité chute > 5% en 1 cycle (60s)
        #    Ignoré si des positions viennent de se fermer ce cycle (faux positif
        #    courant : la marge libérée par un SL/TP n'est pas encore reflétée)
        if hasattr(run, '_prev_equity') and run._prev_equity > 0 and not closed_this_loop:
            equity_drop_pct = (run._prev_equity - balance) / run._prev_equity * 100
            if equity_drop_pct > 5.0:
                msg = (f"Équité: **${run._prev_equity:.2f}** → **${balance:.2f}** "
                       f"(**-{equity_drop_pct:.1f}%** en 1 cycle)")
                print(f"  {C.BRED}[🔴 FLASH CRASH]{C.RST} {equity_drop_pct:.1f}% en 1 cycle")
                send_discord_alert("⚡ Flash Crash Détecté", msg, color=0xe74c3c)
        run._prev_equity = balance

        # 2) Drawdown non-réalisé > 4% de l'équité
        if len(open_positions) > 0 and balance > 0:
            dd_pct = abs(unrealized) / balance * 100
            if unrealized < -balance * 0.04:
                if not getattr(run, '_drawdown_alerted', False):
                    msg = (f"P&L non-réalisé: **${unrealized:.2f}** ({dd_pct:.1f}% de l'équité)\n"
                           f"{len(open_positions)} position(s) ouverte(s)")
                    print(f"  {C.BRED}[📉 DRAWDOWN]{C.RST} {dd_pct:.1f}% de l'équité non-réalisé")
                    send_discord_alert("📉 Drawdown Élevé", msg, color=0xe67e22)
                    run._drawdown_alerted = True
            else:
                run._drawdown_alerted = False

        # 3) Position zombie : ouverte depuis > 6h sans fermeture
        for pos in open_positions:
            opened_at = pos.get('opened_at')
            if opened_at:
                age_h = (time.time() - opened_at) / 3600
                alert_key = f"_zombie_{pos['symbol']}_{pos['id']}"
                if age_h > 6.0 and not getattr(run, alert_key, False):
                    msg = (f"**{pos['symbol']}** ouverte depuis **{age_h:.1f}h** sans TP/SL\n"
                           f"Entrée: ${pos.get('entry', 0):.4f} | SL: ${pos.get('sl', 0):.4f}")
                    print(f"  {C.BYLW}[🧟 ZOMBIE]{C.RST} {pos['symbol']} ouverte depuis {age_h:.1f}h")
                    send_discord_alert("🧟 Position Zombie", msg, color=0xf39c12)
                    setattr(run, alert_key, True)

        # Régénérer le dashboard HTML à chaque cycle
        write_dashboard(wb, open_positions, balance, initial_balance,
                        pnl_total, pnl_week, unrealized, avail_capital,
                        regime, btc_chg, eth_chg)

        elapsed = time.time() - loop_start
        wait    = max(0.0, LOOP_SECONDS - elapsed)
        if wait > 0:
            time.sleep(wait)


if __name__ == "__main__":
    try:
        run()
    except KeyboardInterrupt:
        print(f"\n  {C.BRED}[STOP]{C.RST} Arrêt demandé")
    except Exception as e:
        log_error(f"Crash fatal: {e}")
        traceback.print_exc()
