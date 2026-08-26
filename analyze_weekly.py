"""
analyze_weekly.py — Analyse hebdomadaire du bot par Claude

Usage :
    python analyze_weekly.py

Lit trades.xlsx + lessons.json + .env, envoie les stats à Claude (API Anthropic),
et affiche un rapport avec recommandations d'ajustement des paramètres.

Pré-requis :
    pip install anthropic openpyxl python-dotenv
    ANTHROPIC_API_KEY dans .env (ou variable d'environnement)
"""
import json
import os
from collections import defaultdict
from datetime import datetime, timezone, timedelta

import anthropic
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE   = "trades.xlsx"
LESSONS_FILE = "lessons.json"

# ── Paramètres courants (pour contexte à Claude) ──────────────────────────────
PARAMS = {
    "CAPITAL_PCT":       os.getenv("CAPITAL_PCT",       "0.10"),
    "LEVERAGE":          os.getenv("LEVERAGE",          "5"),
    "ATR_SL_MULT":       "1.5",
    "ATR_TP_MULT":       "4.0",
    "VOL_MULTIPLIER":    os.getenv("VOL_MULTIPLIER",    "1.0"),
    "MAX_DAILY_LOSS_PCT":os.getenv("MAX_DAILY_LOSS_PCT","3.0"),
    "SESSION_START_UTC": os.getenv("SESSION_START_UTC", "0"),
    "SESSION_END_UTC":   os.getenv("SESSION_END_UTC",   "24"),
}


# ── Lecture des trades ─────────────────────────────────────────────────────────
def load_trades() -> list[dict]:
    if not os.path.exists(EXCEL_FILE):
        print(f"[ERREUR] {EXCEL_FILE} introuvable — lance d'abord le bot pour générer les trades.")
        return []
    wb   = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws   = wb["Trades"]
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[14] is None:  # pas de P&L = trade encore ouvert
            continue
        rows.append(dict(zip(hdrs, row)))
    wb.close()
    return rows


# ── Calcul des stats ───────────────────────────────────────────────────────────
def compute_stats(trades: list[dict]) -> dict:
    cutoff_7d = datetime.now(timezone.utc) - timedelta(days=7)

    stats: dict = {
        "total_trades": len(trades),
        "by_symbol": defaultdict(lambda: {"w": 0, "l": 0, "pnl": 0.0}),
        "by_signal": defaultdict(lambda: {"w": 0, "l": 0, "pnl": 0.0}),
        "by_hour":   defaultdict(lambda: {"w": 0, "l": 0, "pnl": 0.0}),
        "by_outcome":{"TP": 0, "SL": 0, "other": 0},
        "week_trades": 0,
        "week_pnl": 0.0,
        "gross_profit": 0.0,
        "gross_loss":   0.0,
        "avg_win": 0.0,
        "avg_loss": 0.0,
        "sl_tp_ratio": 0.0,  # nombre de SLs / TPs
    }

    wins, losses = [], []
    for t in trades:
        pnl    = float(t.get("PnL Net", 0) or 0)
        coin   = str(t.get("Coin", "?"))
        side   = str(t.get("Side", "?"))
        status = str(t.get("Status", ""))
        date   = t.get("Date")

        # Heure d'ouverture
        hour = "?"
        if date:
            try:
                if isinstance(date, str):
                    dt = datetime.fromisoformat(date.replace(" ", "T"))
                else:
                    dt = date
                hour = str(dt.hour)
                # Trades de la semaine
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if dt >= cutoff_7d:
                    stats["week_trades"] += 1
                    stats["week_pnl"]    += pnl
            except Exception:
                pass

        # Par symbole
        s = stats["by_symbol"][coin]
        if pnl > 0:
            s["w"] += 1
            s["pnl"] += pnl
            stats["gross_profit"] += pnl
            wins.append(pnl)
        else:
            s["l"] += 1
            s["pnl"] += pnl
            stats["gross_loss"] += abs(pnl)
            losses.append(pnl)

        # Par signal (BBO / MR / MOM)
        sig_type = "MOM" if "MOM" in str(t.get("Signal", "")) else (
                   "BBO" if "BBO" in str(t.get("Signal", "")) else "MR")
        s2 = stats["by_signal"][f"{side}-{sig_type}"]
        s2["w" if pnl > 0 else "l"] += 1
        s2["pnl"] += pnl

        # Par heure
        h = stats["by_hour"][hour]
        h["w" if pnl > 0 else "l"] += 1
        h["pnl"] += pnl

        # Par outcome
        if "TP" in status.upper():
            stats["by_outcome"]["TP"] += 1
        elif "SL" in status.upper():
            stats["by_outcome"]["SL"] += 1
        else:
            stats["by_outcome"]["other"] += 1

    stats["avg_win"]  = sum(wins)  / len(wins)   if wins   else 0.0
    stats["avg_loss"] = sum(losses)/ len(losses)  if losses else 0.0
    tp_count = stats["by_outcome"]["TP"]
    sl_count = stats["by_outcome"]["SL"]
    stats["sl_tp_ratio"] = round(sl_count / tp_count, 2) if tp_count > 0 else 0.0

    # Convertir defaultdict en dict pour JSON
    stats["by_symbol"] = dict(stats["by_symbol"])
    stats["by_signal"] = dict(stats["by_signal"])
    stats["by_hour"]   = dict(stats["by_hour"])

    return stats


# ── Formatage du prompt ────────────────────────────────────────────────────────
def build_prompt(stats: dict, lessons: dict) -> str:
    pf = round(stats["gross_profit"] / stats["gross_loss"], 2) if stats["gross_loss"] > 0 else 0
    wr = round((stats["by_outcome"]["TP"] / stats["total_trades"] * 100)
               if stats["total_trades"] > 0 else 0, 1)

    sym_table = "\n".join(
        f"  {sym}: {v['w']}W/{v['l']}L  WR={round(v['w']/max(1,v['w']+v['l'])*100)}%  P&L={v['pnl']:+.2f}$"
        for sym, v in sorted(stats["by_symbol"].items(), key=lambda x: x[1]["pnl"])
    )
    sig_table = "\n".join(
        f"  {sig}: {v['w']}W/{v['l']}L  WR={round(v['w']/max(1,v['w']+v['l'])*100)}%  P&L={v['pnl']:+.2f}$"
        for sig, v in sorted(stats["by_signal"].items(), key=lambda x: x[1]["pnl"])
    )
    hour_table = "\n".join(
        f"  {h}h UTC: {v['w']}W/{v['l']}L  P&L={v['pnl']:+.2f}$"
        for h, v in sorted(stats["by_hour"].items(), key=lambda x: int(x[0]) if x[0].isdigit() else 99)
    )
    top_lessons = sorted(lessons.get("buckets", {}).items(),
                         key=lambda x: x[1]["w"] / max(1, x[1]["w"] + x[1]["l"]))[:10]
    lessons_table = "\n".join(
        f"  {k}: WR={round(v['w']/max(1,v['w']+v['l'])*100)}%  {v['w']}W/{v['l']}L  P&L={v['pnl']:+.2f}$"
        for k, v in top_lessons
    )

    return f"""Tu es un expert en trading algorithmique. Analyse les performances hebdomadaires de ce bot de scalping crypto sur Hyperliquid DEX (MAINNET, argent réel) et fournis des recommandations CONCRÈTES d'ajustement de paramètres.

## Stratégie du bot
- Signaux : BBP (Bollinger Band Position) + RSI + Volume sur M15/M5
- Types de signaux : BUY BBO (breakout haussier), BUY MR (mean-reversion bas), SELL MR (mean-reversion haut), SELL MOM (momentum crash)
- SL = 1.5×ATR, TP = 4.0×ATR → ratio R:R théorique = 2.7
- Levier = {PARAMS['LEVERAGE']}× | Capital par trade = {PARAMS['CAPITAL_PCT']} × équité

## Paramètres actuels
{json.dumps(PARAMS, indent=2)}

## Stats globales
- Total trades : {stats['total_trades']}
- Win Rate : {wr}%
- Profit Factor : {pf}
- TP touchés / SL touchés : {stats['by_outcome']['TP']} / {stats['by_outcome']['SL']}  (ratio SL/TP = {stats['sl_tp_ratio']})
- Gain moyen : ${stats['avg_win']:.2f} | Perte moyenne : ${stats['avg_loss']:.2f}
- Semaine ({stats['week_trades']} trades) : P&L = ${stats['week_pnl']:+.2f}

## Performance par symbole
{sym_table}

## Performance par type de signal
{sig_table}

## Performance par heure UTC
{hour_table}

## Patterns les moins performants (leçons du bot)
{lessons_table}

## Ta mission
Fournis une analyse structurée avec :

1. **Diagnostic** : Qu'est-ce qui fonctionne bien ? Qu'est-ce qui ne fonctionne pas ?

2. **Recommandations de paramètres** : Pour CHAQUE paramètre que tu veux modifier, indique EXACTEMENT :
   - Paramètre actuel → valeur recommandée
   - Justification chiffrée (basée sur les stats ci-dessus)

3. **Symboles** : Faut-il exclure certaines cryptos ? Ajouter des filtres spécifiques ?

4. **Horaires** : Y a-t-il des heures UTC à éviter ? Une session à restreindre ?

5. **Risques** : Quels risques vois-tu dans les données actuelles ?

Sois concis et direct. Pas de blabla — des chiffres et des actions concrètes.
"""


# ── Appel à l'API Claude ───────────────────────────────────────────────────────
def analyze_with_claude(prompt: str) -> str:
    client = anthropic.Anthropic()  # lit ANTHROPIC_API_KEY automatiquement

    print("\n  Envoi à Claude pour analyse… (peut prendre 30-60s)\n")

    with client.messages.stream(
        model="claude-sonnet-4-6",  # Sonnet 4.6 : excellent rapport qualité/coût (~$0.03)
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}]
    ) as stream:
        response = stream.get_final_message()

    # Extraire le texte (ignorer les blocs thinking)
    return "\n".join(
        block.text for block in response.content
        if block.type == "text"
    )


# ── Sauvegarde du rapport ──────────────────────────────────────────────────────
def save_report(report: str) -> str:
    ts   = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = f"report_weekly_{ts}.md"
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# Analyse Hebdomadaire — {ts}\n\n")
        f.write(report)
    return path


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  BotShortTrade — Analyse Hebdomadaire par Claude")
    print("=" * 60)

    # Vérification API key
    if not os.getenv("ANTHROPIC_API_KEY"):
        print("\n[ERREUR] ANTHROPIC_API_KEY manquante dans .env")
        print("  Ajoute :  ANTHROPIC_API_KEY=sk-ant-api03-...")
        return

    # Chargement des données
    print("\n  Chargement des trades…")
    trades = load_trades()
    if not trades:
        return
    print(f"  {len(trades)} trades chargés depuis {EXCEL_FILE}")

    print("  Chargement des leçons…")
    lessons = {}
    if os.path.exists(LESSONS_FILE):
        with open(LESSONS_FILE, encoding="utf-8") as f:
            lessons = json.load(f)
        print(f"  {len(lessons.get('buckets', {}))} patterns chargés depuis {LESSONS_FILE}")

    # Calcul des stats
    print("  Calcul des statistiques…")
    stats = compute_stats(trades)

    # Quick summary avant d'appeler Claude
    tp = stats["by_outcome"]["TP"]
    sl = stats["by_outcome"]["SL"]
    wr = round(tp / stats["total_trades"] * 100, 1) if stats["total_trades"] > 0 else 0
    pf = round(stats["gross_profit"] / stats["gross_loss"], 2) if stats["gross_loss"] > 0 else 0
    print(f"\n  📊 Résumé : {stats['total_trades']} trades | WR={wr}% | PF={pf}")
    print(f"     TP={tp} | SL={sl} | Gain moy=${stats['avg_win']:.2f} | Perte moy=${stats['avg_loss']:.2f}")

    # Prompt et appel Claude
    prompt = build_prompt(stats, lessons)
    report = analyze_with_claude(prompt)

    # Affichage
    print("\n" + "=" * 60)
    print("  RAPPORT CLAUDE")
    print("=" * 60)
    print(report)

    # Sauvegarde
    path = save_report(report)
    print(f"\n  ✅ Rapport sauvegardé : {path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
